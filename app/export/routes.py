"""
ME Statistics — Export Routes
================================
Excel (.xlsx) export: staff own data, admin per-user data, admin team summary.
"""

import io
from datetime import datetime
from flask import send_file, request, abort, flash, redirect, url_for
from flask_login import login_required, current_user
from sqlalchemy import func
from app.extensions import db
from app.export import export_bp
from app.models import User, MonthlyReport, Goal, Task
from app.auth.decorators import admin_required, active_required

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

MONTH_NAMES = [
    'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
    'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
]

# Styles
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11) if HAS_OPENPYXL else None
HEADER_FILL = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid') if HAS_OPENPYXL else None
THIN_BORDER = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
) if HAS_OPENPYXL else None


def _style_header(ws, row, ncols):
    """Apply teal header styling to a row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def _build_user_report(wb, user, year):
    """Add sheets for a single user's data."""
    # ── Monthly Reports Sheet ─────────────────────────────────
    ws = wb.create_sheet(title='Monthly Reports')
    headers = ['Month', 'Report Count', 'Target', 'Gap', 'Achievement %']
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    reports = MonthlyReport.query.filter_by(
        user_id=user.id, year=year, approval_status='approved'
    ).order_by(MonthlyReport.month).all()

    report_map = {r.month: r for r in reports}
    ytd_total = 0
    for m in range(1, 13):
        r = report_map.get(m)
        count = r.report_count if r else 0
        target = r.target_snapshot if r else user.monthly_target
        gap = count - target if target > 0 else ''
        ach = f'{round((count / target) * 100, 1)}%' if target > 0 else 'N/A'
        ytd_total += count
        ws.append([MONTH_NAMES[m - 1], count, target, gap, ach])

    ws.append([])
    ws.append(['YTD Total', ytd_total])
    _auto_width(ws)

    # ── Goals Sheet ───────────────────────────────────────────
    ws2 = wb.create_sheet(title='Goals')
    headers2 = ['Title', 'KPI', 'Status', 'Progress %', 'Priority']
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))

    goals = Goal.query.filter_by(
        user_id=user.id, approval_status='approved'
    ).order_by(Goal.created_at.desc()).all()
    for g in goals:
        ws2.append([g.title, g.kpi or '', g.status.replace('_', ' ').title(), g.progress, g.priority.title()])
    _auto_width(ws2)

    # ── Tasks Sheet ───────────────────────────────────────────
    ws3 = wb.create_sheet(title='Tasks')
    headers3 = ['Description', 'Status', 'Progress %', 'Priority']
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))

    tasks = Task.query.filter_by(user_id=user.id).order_by(Task.created_at.desc()).all()
    for t in tasks:
        ws3.append([t.description, t.status.replace('_', ' ').title(), t.progress, t.priority.title()])
    _auto_width(ws3)


def _send_workbook(wb, filename):
    """Convert workbook to bytes and send as download."""
    # Remove default sheet if it exists
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@export_bp.route('/export/my-report')
@login_required
@active_required
def export_my_report():
    """Staff: export own data as Excel."""
    if not HAS_OPENPYXL:
        flash('Export is not available. Please install openpyxl.', 'error')
        return redirect(url_for('dashboard.staff'))

    year = request.args.get('year', datetime.now().year, type=int)
    wb = openpyxl.Workbook()
    _build_user_report(wb, current_user, year)

    filename = f'ME_Report_{current_user.username}_{year}.xlsx'
    return _send_workbook(wb, filename)


@export_bp.route('/export/user/<int:user_id>')
@login_required
@admin_required
def export_user(user_id):
    """Admin: export specific user's data as Excel."""
    if not HAS_OPENPYXL:
        flash('Export is not available. Please install openpyxl.', 'error')
        return redirect(url_for('dashboard.admin'))

    user = User.query.get_or_404(user_id)
    year = request.args.get('year', datetime.now().year, type=int)
    wb = openpyxl.Workbook()
    _build_user_report(wb, user, year)

    filename = f'ME_Report_{user.username}_{year}.xlsx'
    return _send_workbook(wb, filename)


@export_bp.route('/export/team')
@login_required
@admin_required
def export_team():
    """Admin: export team summary as Excel."""
    if not HAS_OPENPYXL:
        flash('Export is not available. Please install openpyxl.', 'error')
        return redirect(url_for('dashboard.admin'))

    year = request.args.get('year', datetime.now().year, type=int)
    wb = openpyxl.Workbook()

    # ── Leaderboard Sheet ─────────────────────────────────────
    ws = wb.create_sheet(title='Leaderboard')
    headers = ['Rank', 'Name', 'YTD Total', 'Avg Monthly', 'Target', 'Achievement %']
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    users = User.query.filter_by(is_active=True, is_approved=True).all()
    board = []
    for u in users:
        ytd = db.session.query(
            func.coalesce(func.sum(MonthlyReport.report_count), 0)
        ).filter(
            MonthlyReport.user_id == u.id,
            MonthlyReport.year == year,
            MonthlyReport.approval_status == 'approved'
        ).scalar()

        months_with_data = MonthlyReport.query.filter_by(
            user_id=u.id, year=year, approval_status='approved'
        ).count()

        avg = round(ytd / months_with_data, 1) if months_with_data > 0 else 0
        target_yearly = u.monthly_target * 12
        ach = f'{round((ytd / target_yearly) * 100, 1)}%' if target_yearly > 0 else 'N/A'

        board.append((u.full_name, ytd, avg, u.monthly_target, ach))

    board.sort(key=lambda x: x[1], reverse=True)
    for rank, (name, ytd, avg, target, ach) in enumerate(board, 1):
        ws.append([rank, name, ytd, avg, target, ach])
    _auto_width(ws)

    # ── Monthly Breakdown Sheet ───────────────────────────────
    ws2 = wb.create_sheet(title='Monthly Breakdown')
    header_row = ['Name'] + MONTH_NAMES + ['YTD']
    ws2.append(header_row)
    _style_header(ws2, 1, len(header_row))

    for u in users:
        reports = MonthlyReport.query.filter_by(
            user_id=u.id, year=year, approval_status='approved'
        ).all()
        row_data = [0] * 12
        for r in reports:
            row_data[r.month - 1] = r.report_count
        ws2.append([u.full_name] + row_data + [sum(row_data)])
    _auto_width(ws2)

    filename = f'ME_Team_Summary_{year}.xlsx'
    return _send_workbook(wb, filename)
