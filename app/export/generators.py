"""Excel and PDF generators for individual and team exports."""

from __future__ import annotations

import calendar
from io import BytesIO
from statistics import mean
from typing import Dict, Iterable, List, Sequence, Tuple

from app.models import Goal, MonthlyReport, Task, User


def _achievement_percentage(report_count: int, target: int) -> str:
    """Return achievement percentage text for one report row.

    Args:
        report_count: Number of processed reports.
        target: Target snapshot value.

    Returns:
        str: Percentage text with one decimal place or `N/A`.

    Side Effects:
        None.

    Raises:
        None.
    """

    if target <= 0:
        return "N/A"
    return f"{(report_count / target) * 100:.1f}%"


def _month_label(month: int) -> str:
    """Return full month name for a month integer.

    Args:
        month: Month number (1-12).

    Returns:
        str: Human-readable month name.

    Side Effects:
        None.

    Raises:
        IndexError: If month is outside 1..12.
    """

    return calendar.month_name[month]


def _summary_rows(reports: Sequence[MonthlyReport]) -> List[Tuple[str, str]]:
    """Build summary table rows for individual exports.

    Args:
        reports: Monthly reports in selected period.

    Returns:
        List[Tuple[str, str]]: Label-value summary rows.

    Side Effects:
        None.

    Raises:
        None.
    """

    if not reports:
        return [
            ("YTD total", "0"),
            ("Average monthly", "0.0"),
            ("Best month", "N/A"),
            ("Target hit rate", "0.0%"),
        ]

    totals = [report.report_count for report in reports]
    ytd_total = sum(totals)
    avg_monthly = mean(totals)
    best_report = max(reports, key=lambda report: report.report_count)

    hit_count = len(
        [
            report
            for report in reports
            if report.target_snapshot > 0 and report.report_count >= report.target_snapshot
        ]
    )
    target_hit_rate = (hit_count / len(reports)) * 100

    return [
        ("YTD total", str(ytd_total)),
        ("Average monthly", f"{avg_monthly:.1f}"),
        ("Best month", f"{_month_label(best_report.month)} ({best_report.report_count})"),
        ("Target hit rate", f"{target_hit_rate:.1f}%"),
    ]


def generate_individual_xlsx(
    user: User,
    reports: Sequence[MonthlyReport],
    goals: Sequence[Goal],
    tasks: Sequence[Task],
) -> bytes:
    """Generate Excel bytes for one user's report/goals/tasks export.

    Args:
        user: Export target user.
        reports: Monthly report rows.
        goals: Goal rows.
        tasks: Task rows.

    Returns:
        bytes: `.xlsx` file bytes.

    Side Effects:
        None.

    Raises:
        ImportError: When openpyxl is unavailable.
    """

    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()

    reports_sheet = workbook.active
    reports_sheet.title = "Monthly Reports"
    reports_headers = [
        "Year",
        "Month",
        "Report Count",
        "Target",
        "Achievement %",
        "MoM Change",
    ]
    reports_sheet.append(reports_headers)

    previous_count = None
    for report in reports:
        mom_change = "N/A"
        if previous_count is not None:
            mom_change = str(report.report_count - previous_count)
        previous_count = report.report_count

        reports_sheet.append(
            [
                report.year,
                _month_label(report.month),
                report.report_count,
                report.target_snapshot,
                _achievement_percentage(report.report_count, report.target_snapshot),
                mom_change,
            ]
        )

    goals_sheet = workbook.create_sheet(title="Goals")
    goals_sheet.append(["Title", "KPI", "Status", "Progress %", "Priority"])
    for goal in goals:
        goals_sheet.append(
            [
                goal.title,
                goal.kpi,
                goal.status.value,
                goal.progress,
                goal.priority.value,
            ]
        )

    tasks_sheet = workbook.create_sheet(title="Tasks")
    tasks_sheet.append(["Description", "Status", "Progress %", "Priority"])
    for task in tasks:
        tasks_sheet.append(
            [
                task.description,
                task.status.value,
                task.progress,
                task.priority.value,
            ]
        )

    summary_sheet = workbook.create_sheet(title="Summary")
    summary_sheet.append(["Staff", user.full_name])
    summary_sheet.append(["Email", user.email])
    summary_sheet.append([])
    summary_sheet.append(["Metric", "Value"])
    for label, value in _summary_rows(reports):
        summary_sheet.append([label, value])

    for sheet in workbook.worksheets:
        sheet["A1"].font = Font(bold=True)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.read()


def generate_team_xlsx(
    users: Sequence[User],
    reports_by_user: Dict[int, Sequence[MonthlyReport]],
    goals_by_user: Dict[int, Sequence[Goal]],
) -> bytes:
    """Generate Excel bytes for team-level admin exports.

    Args:
        users: Staff users included in export.
        reports_by_user: Mapping of user ID to report rows.
        goals_by_user: Mapping of user ID to goal rows.

    Returns:
        bytes: `.xlsx` file bytes.

    Side Effects:
        None.

    Raises:
        ImportError: When openpyxl is unavailable.
    """

    from openpyxl import Workbook

    workbook = Workbook()

    leaderboard_sheet = workbook.active
    leaderboard_sheet.title = "Leaderboard"
    leaderboard_sheet.append(
        ["Rank", "Name", "YTD Total", "Avg Monthly", "Target", "Achievement %"]
    )

    leaderboard_rows: List[Tuple[str, int, float, int, str]] = []
    for user in users:
        reports = list(reports_by_user.get(user.id, []))
        totals = [report.report_count for report in reports]
        ytd_total = sum(totals)
        avg_monthly = mean(totals) if totals else 0.0
        achievement = (
            f"{(avg_monthly / user.monthly_target) * 100:.1f}%"
            if user.monthly_target > 0
            else "N/A"
        )
        leaderboard_rows.append(
            (user.full_name, ytd_total, round(avg_monthly, 1), user.monthly_target, achievement)
        )

    leaderboard_rows.sort(key=lambda item: item[1], reverse=True)
    for rank, row in enumerate(leaderboard_rows, start=1):
        leaderboard_sheet.append([rank, row[0], row[1], row[2], row[3], row[4]])

    breakdown_sheet = workbook.create_sheet(title="Monthly Breakdown")
    breakdown_sheet.append(["Name"] + [calendar.month_abbr[index] for index in range(1, 13)])
    for user in users:
        month_values = [0] * 12
        for report in reports_by_user.get(user.id, []):
            month_values[report.month - 1] = report.report_count
        breakdown_sheet.append([user.full_name] + month_values)

    goals_sheet = workbook.create_sheet(title="Goals Overview")
    goals_sheet.append(["Staff", "Goal count", "% completed"])
    for user in users:
        goals = list(goals_by_user.get(user.id, []))
        goal_count = len(goals)
        completed = len([goal for goal in goals if goal.status.value == "completed"])
        percent_completed = (completed / goal_count) * 100 if goal_count else 0
        goals_sheet.append([user.full_name, goal_count, f"{percent_completed:.1f}%"])

    executive_sheet = workbook.create_sheet(title="Executive Summary")
    department_total = sum(row[1] for row in leaderboard_rows)
    average_total = department_total / len(leaderboard_rows) if leaderboard_rows else 0
    top_name = leaderboard_rows[0][0] if leaderboard_rows else "N/A"
    executive_sheet.append(["Department total", department_total])
    executive_sheet.append(["Department average", round(average_total, 1)])
    executive_sheet.append(["Top performer", top_name])

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.read()


def _pdf_table_style() -> "TableStyle":
    """Return shared table styling for reportlab table widgets.

    Args:
        None.

    Returns:
        TableStyle: Configured style object.

    Side Effects:
        None.

    Raises:
        ImportError: When reportlab is unavailable.
    """

    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D9488")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9FD5CC")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]
    )


def generate_individual_pdf(
    user: User,
    reports: Sequence[MonthlyReport],
    goals: Sequence[Goal],
    tasks: Sequence[Task],
    department_name: str,
) -> bytes:
    """Generate PDF bytes for one user's export packet.

    Args:
        user: Export target user.
        reports: Monthly report rows.
        goals: Goal rows.
        tasks: Task rows.
        department_name: Department display name.

    Returns:
        bytes: `.pdf` file bytes.

    Side Effects:
        None.

    Raises:
        ImportError: When reportlab is unavailable.
    """

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table

    stream = BytesIO()
    document = SimpleDocTemplate(stream, pagesize=A4)
    styles = getSampleStyleSheet()

    story = []
    story.append(Paragraph(f"{department_name} - Individual Export", styles["Title"]))
    story.append(Paragraph(f"Staff: {user.full_name}", styles["Normal"]))
    story.append(Spacer(1, 12))

    reports_table = [["Year", "Month", "Count", "Target", "Achievement %"]]
    for report in reports:
        reports_table.append(
            [
                report.year,
                _month_label(report.month),
                report.report_count,
                report.target_snapshot,
                _achievement_percentage(report.report_count, report.target_snapshot),
            ]
        )
    report_widget = Table(reports_table)
    report_widget.setStyle(_pdf_table_style())
    story.append(report_widget)
    story.append(Spacer(1, 12))

    goals_table = [["Goal", "Status", "Progress", "Priority"]]
    for goal in goals:
        goals_table.append([goal.title, goal.status.value, goal.progress, goal.priority.value])
    goals_widget = Table(goals_table)
    goals_widget.setStyle(_pdf_table_style())
    story.append(goals_widget)
    story.append(Spacer(1, 12))

    tasks_table = [["Task", "Status", "Progress", "Priority"]]
    for task in tasks:
        tasks_table.append([task.description, task.status.value, task.progress, task.priority.value])
    tasks_widget = Table(tasks_table)
    tasks_widget.setStyle(_pdf_table_style())
    story.append(tasks_widget)

    document.build(story)
    stream.seek(0)
    return stream.read()


def generate_team_pdf(
    users: Sequence[User],
    reports_by_user: Dict[int, Sequence[MonthlyReport]],
    department_name: str,
) -> bytes:
    """Generate PDF bytes for team leaderboard and monthly summary.

    Args:
        users: Staff users included in export.
        reports_by_user: Mapping of user ID to monthly reports.
        department_name: Department display name.

    Returns:
        bytes: `.pdf` file bytes.

    Side Effects:
        None.

    Raises:
        ImportError: When reportlab is unavailable.
    """

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table

    stream = BytesIO()
    document = SimpleDocTemplate(stream, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    leaderboard_data = [["Rank", "Name", "YTD Total", "Avg Monthly"]]
    rows: List[Tuple[str, int, float]] = []
    for user in users:
        totals = [report.report_count for report in reports_by_user.get(user.id, [])]
        ytd_total = sum(totals)
        avg_monthly = mean(totals) if totals else 0.0
        rows.append((user.full_name, ytd_total, avg_monthly))

    rows.sort(key=lambda item: item[1], reverse=True)
    for rank, row in enumerate(rows, start=1):
        leaderboard_data.append([rank, row[0], row[1], f"{row[2]:.1f}"])

    month_breakdown = [["Name"] + [calendar.month_abbr[index] for index in range(1, 13)]]
    for user in users:
        month_values = [0] * 12
        for report in reports_by_user.get(user.id, []):
            month_values[report.month - 1] = report.report_count
        month_breakdown.append([user.full_name] + month_values)

    story = [Paragraph(f"{department_name} - Team Export", styles["Title"]), Spacer(1, 12)]

    leaderboard_widget = Table(leaderboard_data)
    leaderboard_widget.setStyle(_pdf_table_style())
    story.append(leaderboard_widget)
    story.append(Spacer(1, 12))

    breakdown_widget = Table(month_breakdown)
    breakdown_widget.setStyle(_pdf_table_style())
    story.append(breakdown_widget)

    document.build(story)
    stream.seek(0)
    return stream.read()
